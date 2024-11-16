#!/usr/bin/env python
# coding: utf-8
import datetime

# In[1]:


import pandas as pd
import re
import numpy as np
import os
import openpyxl
import random
import xlsxwriter
import tkinter as tk
from tkinter import filedialog, font
from subprocess import run
import glob
import pickle
script_dir = os.getcwd()
# 獲取當前工作目錄
os.chdir(script_dir)

#--------------------------------------使用者輸入介面 ----------------------------------------
#--------------------------------------使用者輸入介面函式 ----------------------------------------
import tkinter as tk
from tkinter import filedialog, font
import os
import glob

global heat_part_excel_filepath, TUS_excel_filepath, start_time_excel_filepath, schedule_result_filepath, last_schedule_pkl_filepath
global delete_processing_part

def choose_file(label):
    # 使用filedialog讓使用者選擇檔案
    filepath = filedialog.askopenfilename()
    # 將選擇的檔案路徑顯示在label控件中
    if filepath:
        label.config(text=filepath)

def save_file_paths():
    # 取得三個輸入的excel路徑 並分配到全局變數, 隨後關閉gui介面
    global heat_part_excel_filepath, TUS_excel_filepath, start_time_excel_filepath, schedule_result_filepath, last_schedule_pkl_filepath
    heat_part_excel_filepath = label_heat_process_part.cget("text")
    TUS_excel_filepath = label_TUS_excel.cget("text")
    start_time_excel_filepath = label_start_time_excel.cget("text")
    last_schedule_pkl_filepath = label_last_schedule_pkl.cget("text")
    schedule_result_filepath = label_result_schedule_excel.cget("text")
    root.destroy()

# 定義選擇資料夾的函數
def choose_directory(label):
    directory = filedialog.askdirectory()
    if directory:
        label.config(text=directory)

# 使用者輸入介面主程式
root = tk.Tk()
root.title("自動化排程系統")
root.minsize(width=500, height=500)
root.resizable(width=False, height=False)

# 創建一個Label來顯示熱製程零件Excel檔案路徑
heat_process_part_button = tk.Button(root, text="1.輸入: 「熱處理製程零件Excel」",height=3, width=30,font=("標楷體",10, "bold"), command=lambda: choose_file(label_heat_process_part))
heat_process_part_button.pack(pady=20)
label_heat_process_part = tk.Label(root, width=50, bg='white')
label_heat_process_part.pack(pady=10)


# 創建一個Label來顯示TUS_excel檔案路徑
TUS_excel_button = tk.Button(root, text="2.輸入: 「真空爐校驗表單Excel」",height=3, width=30,font=("標楷體", 10, "bold"), command=lambda: choose_file(label_TUS_excel))
TUS_excel_button.pack(pady=30)
label_TUS_excel = tk.Label(root, width=50, bg='white')
label_TUS_excel.pack(pady=10)

# 創建一個Label來顯示start_time_excel檔案路徑
start_time_excel_button = tk.Button(root, text="""3.輸入: 「程式開始時第一批零件
          進爐時間設定Excel」""",height=3, width=30,font=("標楷體", 10, "bold"), command=lambda: choose_file(label_start_time_excel),anchor='w',  # 左對齊
    justify='left' )
start_time_excel_button.pack(pady=20)
label_start_time_excel = tk.Label(root, width=50, bg='white')
label_start_time_excel.pack(pady=10)

# 創建一個Label來顯示上一個排程的pkl檔案路徑
last_schedule_pkl_button = tk.Button(root, text="4.輸入: 「前一次排程數據Pickle」",height=3, width=30,font=("標楷體", 10, "bold"), command=lambda: choose_file(label_last_schedule_pkl))
last_schedule_pkl_button.pack(pady=20)
label_last_schedule_pkl = tk.Label(root, width=50, bg='white')
label_last_schedule_pkl.pack(pady=10)

# 創建一個Label來顯示儲存檔案路徑
result_schedule_excel_button = tk.Button(root, text="5.輸出: 「熱處理排程表單Excel」",height=3, width=30,font=("標楷體", 10, "bold"), command=lambda: choose_directory(label_result_schedule_excel))
result_schedule_excel_button.pack(pady=20)
label_result_schedule_excel = tk.Label(root, width=50, bg='white')
label_result_schedule_excel.pack(pady=10)

submit_button = tk.Button(root, text="執行系統", width=30,font=("標楷體", 10, "bold"),command=save_file_paths)
submit_button.pack(pady=20)

root.mainloop()
#--------------------------------------使用者輸入介面----------------------------------------

#--------------------------------------step 3 ----------------------------------------
#--------------------------------------step 3 函式---------------------------------------
# workbook_buffer為一維列表 儲存excel資料本身的列表
# sheet_names_buffer為二維列表 儲存excel資料工作表名稱的列表
# sheet_buffer 一維列表 (存在對應excel的第一個(最左邊)的工作表)
def excel_init_setup(buffer):
    workbook_buffer = []
    sheet_names_buffer = []
    sheet_buffer = []
    for i in buffer:
        # 開啟 Excel 檔案
        workbook = openpyxl.load_workbook(i)
        workbook_buffer.append(workbook)
        # 取得工作表名稱列表
        sheet_names = workbook.sheetnames
        sheet_names_buffer.append(sheet_names)
        # 開啟工作表
        sheet = workbook[sheet_names[0]]
        sheet_buffer.append(sheet)
    return workbook_buffer, sheet_names_buffer, sheet_buffer
# 3.5 開始 擷取熱處理站的零件欄位資訊與技術文件的關鍵資訊並輸入到新Excel試算表1
# 回傳指定tag在sheet工作表中的row以及col位置
def get_tag_position(sheet, tag):
    row = -1
    col = -1
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row = i, column = j).value == tag:
                row = i
                col = j
                break
        if row != -1 and col != -1:
            break
    return row, col

# 將工作表(sheet)中的指定tag下列所有的值儲存至list並回傳
def get_tag_all_(sheet, tag, row = -1, col = -1):
    buffer = []
    # 如果沒有初始位置則從頭開始找
    if row == -1 or col == -1:
        row, col = get_tag_position(sheet, tag)

    for i in range(row+1, sheet.max_row+1):
        buffer.append(sheet.cell(i, col).value)

    return buffer

# 回傳從最左邊的tag開始的所有tag都有的list。
def get_all_tag_(sheet, tag):
    buffer = []
    row, col = get_tag_position(sheet, tag)
    for i in range(col, sheet.max_column):
        buffer.append(sheet.cell(row, i).value)
        print(sheet.cell(row, i).value)
    return buffer
# 3.5 結束
# 3.3 開始,讀取技術文件資料夾並儲存所有檔案名稱
# 將指定資料夾路徑的所有指定xlsx副檔名的檔案名稱放入一個列表並返回
def list_files_in_directory(directory_path, file_extension=".xlsx"):
    files_list = [f for f in os.listdir(directory_path) if f.endswith(file_extension)]
    return files_list
# 3.3 結束
# 3.5 開始 擷取熱處理站的零件欄位資訊與技術文件的關鍵資訊並輸入到新Excel試算表1
# 將s2_excel中的資料填入s1_excel
# s1_sheet s2_sheet 工作表
# s1_row s2_row 固定值
# s1_col s2_col 列表
def excel_replace(s1_sheet, s1_row, s1_col, s2_sheet, s2_row, s2_col):
    for i in range(len(s1_col)):
        s1_sheet.cell(s1_row, s1_col[i]).value = s2_sheet.cell(s2_row, s2_col[i]).value
# tag_list接受一維陣列 建立一個Excel 欄位名稱有熱處理站Excel的欄位 與 技術文件關鍵資訊的名稱
def excel_tag_init(excel_name = 'new_excel.xlsx', table_name = 'init_sheet', tag_list = None):
    if tag_list == None:
        print("row or col or tag_list is empty...")
        return
    # 初始化一個新的excel以及兩個工作表
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb['Sheet'])
    new_wb.create_sheet(f'{table_name}')
    new_wb.create_sheet("被過濾的製程")
    s1 = new_wb[f'{table_name}']
    s2 = new_wb['被過濾的製程']
    s1_col = 1
    s2_col = 1
    fixed_width = [21.5, 10, 105, 7, 28, 50, 50, 20, 30, 30, 30, 30, 30, 30, 30, 30]
    fixed_pos = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
    #設置'A', 'B', 'C'...等欄位寬度
    for i in range(len(fixed_pos)):
        s1.column_dimensions[fixed_pos[i]].width = fixed_width[i]
        s2.column_dimensions[fixed_pos[i]].width = fixed_width[i]
    for i in tag_list:
        s1.cell(1, s1_col).value = i
        s1_col+=1
    for i in tag_list:
        s2.cell(1, s2_col).value = i
        s2_col+=1
    new_wb.save(excel_name)

# init.xlsx的Excel(技術文件)初始化
#傳入 excel 像是技術文件excel 和欄位名稱
#可得到打開後的技術文件excel和類別, 時間, 真空爐爐號 ,數量限制, 溫度位置
def init_xlsx_init(source_excel, source_tag_name):
    workbook = openpyxl.load_workbook(source_excel)
    sheet_names = workbook.sheetnames
    source_sheet = workbook[sheet_names[0]]
    source_sheet_row = []
    source_sheet_col = []
    #取得技術文件 類別, 時間, 真空爐爐號 ,數量限制, 溫度位置
    for i in source_tag_name:
        temp_row , temp_col = get_tag_position(source_sheet, i)
        source_sheet_row.append(temp_row)
        source_sheet_col.append(temp_col)
        #print(f"temp_row :{temp_row}. temp_col :{temp_col}")
    return source_sheet, source_sheet_row, source_sheet_col

# 根據類別的類型，填入類別以及時間欄位
# step 3 output_excel 試算表1 , 該零件技術文件Excel, 第幾個零件 , 技術文件關鍵資訊的名稱與step 3 output_excel欄位位置
def excel_relate_write(target_sheet, source_excel, row, source_tag_name, source_tag_number):
    source_sheet, source_sheet_row, source_sheet_col = init_xlsx_init(source_excel, source_tag_name)
    # 取得零件技術文件的所有製程類別
    class_buffer = get_tag_all_(source_sheet, source_tag_name[0], source_sheet_row[0]+1, source_sheet_col[0])
    # 取得零件技術文件的所有時間_第一部分
    time_buffer = get_tag_all_(source_sheet, source_tag_name[1], source_sheet_row[0]+1, source_sheet_col[1]+1)
    # 處理類別欄，並比對類別對時間做相加
    string_buffer = ""
    sum = 0
    # 根據類別額外計算工時 ，同時填入Excel的類別欄位
    for j in class_buffer:
        if j != None:
            string_buffer = string_buffer + str(j) + ','
            # 硬銲製程 額外加上6小時工時
            if j == "硬銲":
                sum += 360
            # 非硬銲製程 額外加上5小時工時
            else :
                sum += 300
    target_sheet.cell(row, source_tag_number[0]).value = string_buffer.strip(',')
    # 處理時間欄，所有時間_第一部分加總，未滿1小時的部分要無條件進位，同時填入Excel的時間欄位
    for j in time_buffer:
        if j != None:
            sum = sum + int(j)
    if sum % 60 !=0:
        sum = (sum // 60) + 1
    else:
        sum = sum // 60
    target_sheet.cell(row, source_tag_number[1]).value = str(sum) + "hr"
# 處理真空爐爐號欄，指定的excel填入某個零件的真空爐號資訊
# 傳入step 3 output_excel 試算表1 , 該零件技術文件Excel, 第幾個零件 , 技術文件關鍵資訊的名稱與step 3 output_excel欄位位置
def Stove_judge(target_sheet, source_excel, row, source_tag_name, source_tag_number):
    source_sheet, source_sheet_row, source_sheet_col = init_xlsx_init(source_excel, source_tag_name)
    tag_length = len(source_tag_name) - 1
    buffer = []
    string_buffer = ""
    stove_to_number = ["A0064232(32001749)", "A0041090(32001661)", "32003075-1", "32013259", "32018276", "32022219", "32025349", "32025350"]
    # 得到 所有的真空爐數量限制
    buffer = get_tag_all_(source_sheet, source_tag_name[tag_length], source_sheet_row[0]+1, source_sheet_col[tag_length])
    # 對每真空爐數量限制
    for i in range(len(buffer)):
        target_sheet.cell(row, source_tag_number[tag_length] + i).value = buffer[i]
        # 若數量限制不是N 則記錄真空爐爐號
        if buffer[i] != 'N':
            string_buffer = string_buffer + stove_to_number[i] + ','
        # 填入 數量限制到step 3 output_excel 真空爐爐號欄位
        target_sheet.cell(row, source_tag_number[tag_length] + i).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        # 填入 記錄的真空爐爐號到step 3 output_excel 真空爐爐號欄位
    target_sheet.cell(row, source_tag_number[tag_length-1]).value = string_buffer.strip(',')

# 處理溫度欄，找到製程階段的溫度 指定的excel填入某個零件的溫度資訊
# 傳入step 3 output_excel 試算表1 , 該零件技術文件Excel, 第幾個零件 , 技術文件關鍵資訊的名稱與step 3 output_excel欄位位置
def temp_judge(target_sheet, source_excel, row, source_tag_name, source_tag_number):
    source_sheet, source_sheet_row, source_sheet_col = init_xlsx_init(source_excel, source_tag_name)
    tag_length = len(source_tag_name)
    for i in range(0, tag_length):
        # 取得技術文件所有的溫度
        temp_buffer = get_tag_all_(source_sheet, source_tag_name[i], source_sheet_row[0]+1, source_sheet_col[i])
        # 取得技術文件所有的階段
        class_buffer = get_tag_all_(source_sheet, "階段", source_sheet_row[0]+1, source_sheet_col[i]-1)
        string_buffer = ""
        # 如果階段為製程 紀錄該溫度資訊
        for j in range(len(temp_buffer)):
            if temp_buffer[j] != None and class_buffer[j] == "製程":
                string_buffer = string_buffer + temp_buffer[j] + ','
        # 把溫度資訊填入step 3 output_excel的製程溫度需求欄位
        target_sheet.cell(row, source_tag_number[i]).value = string_buffer.strip(',')
# 3.5 結束
#丟入 buffer中的第一個excel 例如0130熱處理 ,技術文件名稱列表、要存放轉站正確零件的step3_output_excel(有兩個工作表)、熱處理站零件Excel欄位名稱 與 技術文件關鍵資訊名稱
def Step3_Classification(sheet, buffer, Excel_name, sheet_tag_name, sheet_tag_number, source_tag_name, source_tag_number):
    # 3.5 3.6開始 擷取熱處理站的零件欄位資訊與技術文件的關鍵資訊並輸入到新Excel試算表1
    workbook = openpyxl.load_workbook(Excel_name)
    sheet_names = workbook.sheetnames
    # step3_output_excel 試算表1、2
    s1 = workbook[sheet_names[0]]
    s2 = workbook[sheet_names[1]]
    s1_row = 2
    s2_row = 2
    sheet_row = []
    sheet_col = []
    # 回傳如0130熱處理excel 的件號、工作號碼、數量、出站日期之excel表格位置
    for i in sheet_tag_name:
        temp_row , temp_col = get_tag_position(sheet, i)
        sheet_row.append(temp_row)
        sheet_col.append(temp_col)
    # 3.5 3.6結束
    # 3.4 開始,確認零件是否有對應技術文件Excel
    # 每一筆資料 用件號和步序去和技術文件列表中的件號步序名稱對比
    # 如果有 就把熱處理站零件Excel的零件四個欄位的資訊都填入填入Step 3 output excel 的第一個試算表
    # 並且 填入經過處理的技術文件關鍵資訊,沒有就填入試算表2
    #對每個零件 1,2,...n
    for i in range(sheet_row[0]+1, sheet.max_row+1):
        found_element = None
        #buffer是技術文件檔名清單
        for j in buffer:
            #如果零件件號_工作號碼符合技術文件excel檔名  就零件的件號、工作號碼、數量、出站日期與技術文件關鍵資訊(處理過)把寫入s1
            if f"{sheet.cell(i, sheet_col[0]).value}_{sheet.cell(i, sheet_col[1]).value}" in j:
            # 3.4 結束
                #取出某個零件excel檔名
            # 3.5 開始 擷取熱處理站的零件欄位資訊與技術文件的關鍵資訊並輸入到新Excel試算表1
                found_element = j
                # 第一個階段填入
                # 把熱處理站的零件Excel 一筆資料中的件號、工作號碼、數量、出站日期的資料填入new_excel件號、工作號馬、數量、出站日期
                excel_replace(s1, s1_row, sheet_tag_number, sheet, i, sheet_col)
                # 第二個階段填入
                # 找到該零件的技術文件Excel 並傳入函式中
                excel_path = f"技術文件Excel/{found_element}"
                # 將一筆零件的類別、時間資訊、經過處理填入s1
                excel_relate_write(s1, excel_path, s1_row, source_tag_name[0], source_tag_number[0])
                # 將一筆零件的真空爐號、數量限制經過處理填入s1
                Stove_judge(s1, excel_path, s1_row, source_tag_name[1], source_tag_number[1])
                # 將一筆零件的溫度經過處理填入s1
                temp_judge(s1, excel_path, s1_row, source_tag_name[2], source_tag_number[2])
                s1_row+=1
                break
            # 3.5 結束
        # 3.6 開始
        # 沒有找到則寫入s2
        if found_element == None:
            excel_replace(s2, s2_row, sheet_tag_number, sheet, i, sheet_col)
            s2_row+=1
        # 3.6 結束
    # 3.7 開始 輸出Step 3 Excel結果
    workbook.save(Excel_name)
    # 3.7 結束


#--------------------------------------step 3 主程式----------------------------------------
script_dir = os.getcwd()
# 欲使用的Excel檔。

# 3.1 開始,熱處理站內零件Excel去除第一列空白資料
part_in_heattreatment_df = pd.read_excel(heat_part_excel_filepath)
# 資料前處理
# 由於excel第一列是空的 因此欄位會變成unname 所以將第二列的資列轉換為dataframe的欄位
part_in_heattreatment_df.columns = part_in_heattreatment_df.iloc[0]
part_in_heattreatment_df = part_in_heattreatment_df[1:]
part_in_heattreatment_df.reset_index(drop=True, inplace=True)
part_in_heattreatment_df.to_excel(os.path.join(script_dir, '程式間接輸出/熱處理_preprocessing.xlsx'),index=False)
# 3.1 結束

# 3.2 開始,建立新Excel
buffer = [os.path.join(script_dir, '程式間接輸出/熱處理_preprocessing.xlsx'), os.path.join(script_dir, 'init.xlsx')]
# 回傳三個變數 -> 儲存Excel的變數 , 儲存Excel所有sheet名稱的變數 , 儲存Excel第一個(最左)sheet的變數
workbook_buffer, sheet_names_buffer, sheet_buffer = excel_init_setup(buffer)
# 返回workbook_buffer 裡面元素為每一個excel 的開啟
# 返回sheet_names_buffer 裡面有多個列表 每個列表代表 一個excel 列表元素代表excel的工作表名稱
# 返回sheet_buffer 開啟excel(熱處理_preprocessing)最左邊的工作表
# Tag所有名稱 - 最後的excel需要的所有column名稱
# tag所有名稱 輸入 nwe_excel 和某個工作表名稱、欄位列表 輸出已經添加全新欄位名稱調整表格寬度的兩個表格
buffer = ["件號", "工作號碼", "真空爐爐號", "數量", "計畫出站日期", "製程溫度需求", "時間", "類別", "A0064232(32001749)數量限制", "A0041090(32001661)數量限制", "32003075-1數量限制", "32013259數量限制", "32018276數量限制", "32022219數量限制", "32025349數量限制", "32025350數量限制"]
# 透過buffer的名稱先生成一個指定名稱為欄位的step3_output_Excel檔案
# Excel欄位為"件號", "工作號碼", "真空爐爐號", "數量", "計畫出站日期", "製程溫度需求", "時間", "類別", "A0064232(32001749)數量限制", "A0041090(32001661)數量限制", "32003075-1數量限制", "32013259數量限制", "32018276數量限制", "32022219數量限制", "32025349數量限制", "32025350數量限制"
excel_tag_init(os.path.join(script_dir, '程式間接輸出/step3_output_excel.xlsx'), sheet_names_buffer[0][0], buffer)
# 3.2 結束

# 3.3 開始,讀取技術文件，並儲存資料夾中的檔名
# 技術文件的資料夾位置
directory_path = '技術文件Excel'
# 列出資料夾中的指定格式的檔案
xlsx_files = list_files_in_directory(directory_path, ".xlsx")
# Step 3.3 結束
# 3.4 3.5 3.6 3.7 開始, 都會用到這些代碼
# Step3_Classification需要的變數
# sheet_buffer[0] 對應 tag所有名稱 的index位置設定

tag_name = ["件號", "步序", "數量", "計畫出站日期"]
tag_number = [1,2,4,5] # 對照最終xlsx的各個tag的位置

## 技術文件的 對應 tag所有名稱 的index位置設定

source_tag_name = [["類別", "時間"], ["真空爐爐號", "數量限制"], ["溫度"]]
source_tag_number = [[8, 7], [3,9], [6]]

# 丟入 buffer中的第一個excel 例如0130熱處理 ,技術文件名稱列表、要存放的step3_output_excel(有兩個工作表)、熱處理站零件Excel欄位名稱 與 技術文件關鍵資訊名稱
Step3_Classification(sheet_buffer[0], xlsx_files, os.path.join(script_dir, '程式間接輸出/step3_output_excel.xlsx'), tag_name, tag_number, source_tag_name, source_tag_number)
# 可以輸出step3_output_excel 裡面的兩個表格:分別為1. 轉站正確的零件資料 2.轉站錯誤的零件資料
# 3.4 3.5 3.6 3.7 結束
#--------------------------------------step 3  ----------------------------------------
#--------------------------------------step 4 ----------------------------------------
# 4.1 開始,輸入Step 3 Excel結果
# 輸入Step3產生的零件資料
step4_input_part_df = pd.read_excel(os.path.join(script_dir, '程式間接輸出/step3_output_excel.xlsx'), converters={'工作號碼': str})
#去除計畫出站日期的"上午"
step4_input_part_df['計畫出站日期'] = step4_input_part_df['計畫出站日期'].str.replace('上午 ', '')
#計畫出站日期資料型態從字串轉為timestamp
step4_input_part_df['計畫出站日期'] = pd.to_datetime(step4_input_part_df['計畫出站日期'])
# 4.1 結束
# 4.2 開始 把Step 3 Excel每一批相同件號與工作號碼的零件分成一組並把組內的零件數量加總，同時取最早預計出站日期
# 定義要使用的欄位名稱列表
columns = ['件號', '工作號碼', '真空爐爐號', '數量', '計畫出站日期',
           '製程溫度需求', '時間', '類別', 'A0064232(32001749)數量限制',
           'A0041090(32001661)數量限制', '32003075-1數量限制', '32013259數量限制',
           '32018276數量限制', '32022219數量限制', '32025349數量限制', '32025350數量限制']
# 將step4_input_part_df的excel 根據件號和步序進行分組(件號和步序都要相同才算是一組)  再把數量加總
# 並把組內的最早出站日期抓出來
step4_input_part_df = step4_input_part_df.groupby(['件號', '工作號碼']).agg({
    '數量': 'sum',  # 合併數量
    '計畫出站日期': 'min',  # 選擇最早的計畫出站日期
    # 對於其他欄位，選擇分組後第一條記錄的值
    '真空爐爐號': 'first',
    '製程溫度需求': 'first',
    '時間': 'first',
    '類別': 'first',
    'A0064232(32001749)數量限制': 'first',
    'A0041090(32001661)數量限制': 'first',
    '32003075-1數量限制': 'first',
    '32013259數量限制': 'first',
    '32018276數量限制': 'first',
    '32022219數量限制': 'first',
    '32025349數量限制': 'first',
    '32025350數量限制': 'first'
}).reset_index()
# 4.2 結束
# 4.3 開始
step4_input_part_df.to_excel(os.path.join(script_dir, '程式間接輸出/step4_output_excel.xlsx'),index=False)
# 4.3 結束
#--------------------------------------step 4 ----------------------------------------
#--------------------------------------step 5 ----------------------------------------
# 5.1 開始,輸入Step 4 Excel結果
step5_input_part_df = pd.read_excel(os.path.join(script_dir, '程式間接輸出/step4_output_excel.xlsx'), converters={'工作號碼': str})
# 5.1 結束
# 5.2 開始,所有零件使用計畫出站日期進行升序排序(越早的計畫出站日期越前面)
# 排序 升序(從以前排到現在)
step5_input_part_df.sort_values(by='計畫出站日期',inplace=True)
# 5.2 結束
# 5.3 開始,輸出Step 5 Excel結果
step5_input_part_df.to_excel(os.path.join(script_dir, '程式間接輸出/step5_output_excel.xlsx'),index=False)
# 5.3 結束
#--------------------------------------step 5 ----------------------------------------

#--------------------------------------step 6 ----------------------------------------
#--------------------------------------step 6 函式區----------------------------------------
#6.2 開始,均溫校驗表前處理
#輸入 溫度校驗表excel路徑，作資料前處理 返回欄位名:真空爐號 資料:真空爐均溫校驗點的 dataframe
def TUS_get_temp_point_df(TUS_excel_path):
    TUS_excel_df = pd.read_excel(TUS_excel_path,header=1)
    # 找出TUS point (℃) & Furnace class 和 校驗日期 的index位置
    start_index = TUS_excel_df[TUS_excel_df["爐號 Furnace No."]=="TUS point (℃) & Furnace class"].index[0]
    end_index = TUS_excel_df[TUS_excel_df['爐號 Furnace No.'] == "校驗日期"].index[0]
    # 擷取TUS point (℃) & Furnace class 和 校驗日期 資訊
    TUS_excel_df = TUS_excel_df.iloc[start_index:end_index]
    # 去除欄位 爐號 Furnace No.
    TUS_excel_df.drop(columns="爐號 Furnace No.",inplace=True)
    # 把均溫校驗表的真空爐改名 統一叫做A0064232(32001749) A0041090(32001661)
    TUS_excel_df.rename(columns={"A0064232 / 32001749": "A0064232(32001749)", "A0041090 / 32001661": "A0041090(32001661)"}, inplace=True)
    TUS_excel_df = TUS_excel_df.reset_index(drop=True)
    # 7/20更新 TUS Point的欄位 資料型態不一致(有int str) 統一轉成str
    TUS_excel_df.columns = TUS_excel_df.columns.astype(str)
    return TUS_excel_df
#6.2 結束,均溫校驗表前處理
#6.4 開始 將均溫校驗表去除故障的真空爐校驗資訊，並把校驗溫度點作處理
#傳入 某個真空爐欄位和均溫校驗點數量 返回二維列表 列表的元素是某個真空盧經過去除NAN ±符號分割和去除不符合規則的均溫校驗點
def TUS_temp_point_split(TUS_vacuum_temp_series, TUS_temp_point_num):
    result = []
    pattern = re.compile(r'^\d+±\d+$')  #建立正則表達式 建立規則 只需要±前後都有整數的字串 為了要去除(300±40Remove)
    # 對vacuum temp series 的每個數值進行 ±符號的分割 和去除nan和不符合規則的數值
    for temp in range(TUS_temp_point_num):
        current_element = TUS_vacuum_temp_series.iloc[temp]
        #去除空格，為了符合正則表達式的規則
        current_element = str(current_element).strip()
        # 檢查元素是否為NaN
        if pd.isna(current_element):
            continue  # 如果是NaN，略過當前循環的剩餘部分
        if not pattern.match(current_element):
            continue  # 如果不符合規則，就略過
        # 分割每個數據點 為列表 [溫度點,容差]
        split_point = current_element.split('±')
        # 分割後 把±前後都轉成整數
        result.append(list(map(int, split_point)))
    return result
#6.4 結束
#6.7 開始 對零件的製程溫度進行資料處理
# 傳入一個零件的製程溫度需求 例如:300±15 輸出二維列表 按照± 分開成300 ,15 得到[[300,15]]
def part_temp_split(part_temp):
    result = []
    part_temp_list = part_temp.split(",")
    for part in part_temp_list:
        part_temp_point_range = part.split("±")
        part_temp_point_range = list(map(int,part_temp_point_range))
        result.append(part_temp_point_range)
    return result
# 6.7結束
#6.4 開始 將均溫校驗表去除故障的真空爐校驗資訊，並把校驗溫度點作處理
#傳入 真空爐清單和校驗表 TUS Point 的row 數量 返回一個字典 key是真空爐號 value 是真空爐對應的TUS point校驗溫度點清單
def vaccum_temp_build_dict(vac_list,TUS_point_row):
    TUS_vacuum_temp_list = []
    result_dict = None
    # 建立三維列表 不同真空爐的不同校驗溫度
    for vac in vac_list:
        TUS_vacuum_temp_list.append(TUS_temp_point_split(TUS_vacuum_temp_df[vac],TUS_point_row))
    # 建立字典 一個真空爐對應多個溫度校驗點[溫度點,容差]
    for vac_temp in TUS_vacuum_temp_list:
        result_dict = {str(key): value for key, value in zip(vac_list,TUS_vacuum_temp_list )}
    return  result_dict
#6.4 結束
# 6.8 開始 零件的所有溫度需求與每個真空爐校驗溫度比較，篩選並紀錄符合條件的爐號
#傳入 製程溫度二維列表(已經將±分離) 和 零件真空爐爐號列表和校驗表溫度資料(字典) 可以返回一個列表 裡面是零件的最終可進入的真空爐爐號
def check_temperature_range(process_temp_list, part_vac_list,vacuum_temp_dict_input):
    vaccum_result = []
    for vac in part_vac_list:
# 前提字典的真空爐格式和零件excel的真空爐爐號格式相同
# TUS_point_detection_list 列表元素為一個溫度點和容差
        process_temp_True_False_result = []
        #某個真空爐的均溫校驗點列表
        TUS_point_detection_list = vacuum_temp_dict_input[vac]
        #從零件的製程溫度需求一一取出來
        for process_temp in process_temp_list:
            process_temp_center = process_temp[0]
            process_temp_tolerance = process_temp[1]
            process_temp_fit_center = False
            #一開始會看製程溫度中心點是否已經比較過"等於校驗表溫度中心點"
            if not process_temp_fit_center:
                #將每一個校驗溫度的都進行
                #如果製程溫度中心點有等於校驗溫度中心點 則比較製程溫度容差是否大於校驗溫度容差 如果有 則標記符合中心點True 並加入結果True
                # 並進行下一輪製程溫度的比較,如果沒有 則依樣進行下一輪零件的比較 但是符合中心點為False
                for TUS_point in TUS_point_detection_list:
                    tus_point_center = TUS_point[0]
                    tus_point_tolerance = TUS_point[1]
                    if process_temp_center == tus_point_center:
                        if process_temp_tolerance >= tus_point_tolerance:
                            process_temp_fit_center = True
                            process_temp_True_False_result.append(True)
                            break
                        else:
                            process_temp_fit_center = True
                            process_temp_True_False_result.append(False)
                            break
            #比較過中心點之後 如果沒有製程溫度點等於校驗中心點 則看有沒有在某兩個校驗中心點的區間內
            if not process_temp_fit_center:
                #先測試製程溫度是否低於校驗表最低溫度或者是否高於最高溫度 如果有此製程溫度就不被滿足
                if process_temp_center < TUS_point_detection_list[0][0]:
                        process_temp_True_False_result.append(False)
                elif process_temp_center > TUS_point_detection_list[len(TUS_point_detection_list)-1][0]:
                        process_temp_True_False_result.append(False)
                    #找出校驗溫度列表的所有溫度點和容差
                for index in range(len(TUS_point_detection_list)-1):
                    #零件製程溫度去比較每兩個校驗溫度點 如果找到在某個區間 則比較容差(兩校驗溫度點使用較大的容差取比較)
                    if TUS_point_detection_list[index][0] < process_temp_center < TUS_point_detection_list[index+1][0]:
                        TUS_point_max_tolerance = max(TUS_point_detection_list[index][1],TUS_point_detection_list[index+1][1])
                        #製程溫度容差如果比較大，代表這個真空爐的校驗溫度滿足這個製程溫度點
                        if process_temp_tolerance >= TUS_point_max_tolerance:
                            process_temp_True_False_result.append(True)
                        else:
                            process_temp_True_False_result.append(False)

        #最後看結果是否都是TRUE，來決定要步要進這個真空爐，只有一個零件溫度是false 就代表不符合校驗溫度，則不可進此真空爐，全部TRUE才可以
        if all(process_temp_True_False_result):
            vaccum_result.append(vac)
    return  vaccum_result
#6.8 結束
#6.3 開始,均溫校驗表中找到故障的真空爐爐號
#輸入 溫度校驗表excel路徑 返回 所有機障中的熱處理爐爐號
def TUS_get_m(TUS_excel_path):
        TUS_excel_df = pd.read_excel(TUS_excel_path,header=1)
        max_row_index = TUS_excel_df.shape[0] - 1
        start_index = TUS_excel_df[TUS_excel_df["爐號 Furnace No."]=="TUS point (℃) & Furnace class"].index[0]
        TUS_excel_df = TUS_excel_df.iloc[start_index:max_row_index]
        TUS_excel_df = TUS_excel_df.reset_index(drop=True)

        TUS_excel_df.rename(columns={"A0064232 / 32001749": "A0064232(32001749)", "A0041090 / 32001661": "A0041090(32001661)"}, inplace=True)
        # # 7/20更新 TUS Point的欄位 資料型態不一致(有int str) 統一轉成str
        TUS_excel_df.columns = TUS_excel_df.columns.astype(str)
        search_string = "機障中"
        # 儲存機障中的熱處理爐
        result_list = []
        for column_name in TUS_excel_df.columns:
            # 取得該欄位的索引位置
            column_index = TUS_excel_df.columns.get_loc(column_name)
            # 選取指定欄位下方的所有資訊
            selected_data = TUS_excel_df.iloc[:, column_index]
            # 搜尋 機障中 在 各欄所有資訊中的存在結果
            result = selected_data.str.contains(search_string)
            # 從result中判斷True是否存在
            exists = result.any()
            if exists == True:
                result_list.append(column_name)
        print("機障中")
        for i in result_list:
            print(i)
        return result_list
#Step6.3結束
#--------------------------------------step 6 主程式 ----------------------------------------
#6.1 開始
# 讀取Step5 輸出的零件資料
step6_input_excel_path_directory = os.path.join(script_dir, '程式間接輸出/step5_output_excel.xlsx')
#6.1 結束
#6.2 開始,讀取均溫校驗表，接著均溫校驗表前處理
step6_input_excel_df = pd.read_excel(step6_input_excel_path_directory, converters={'工作號碼': str})

# 輸入真空爐校驗表 保留真空爐爐號和均溫校驗點TUS Point 剔除其他資訊
TUS_vacuum_temp_df = TUS_get_temp_point_df(TUS_excel_filepath)

#6.2 結束
#6.3 開始,均溫校驗表中找到故障的真空爐爐號
#從真空爐均溫校驗表找到TUS Point列之中有"機障中"的真空爐爐號
bad_stove = TUS_get_m(TUS_excel_filepath)
#6.3 結束
#6.4 開始 將均溫校驗表去除故障的真空爐校驗資訊，並把校驗溫度點作處理
#依照校驗表的真空爐號 列出目前所有的真空爐號
vacuum_list = TUS_vacuum_temp_df.columns.tolist()
for vac in vacuum_list:
    print(type(vac))
# 使用列表推導式過濾出不在 bad_stove 中的元素
vacuum_list = [x for x in vacuum_list if x not in bad_stove]
# 產生字典 真空爐爐號: 溫度校驗點
TUS_point_rownum = TUS_vacuum_temp_df.shape[0]
vacuum_TUS_dict = vaccum_temp_build_dict(vacuum_list,TUS_point_rownum)
#6.4 結束

# 將零件的所有製程溫度需求 和每一個真空爐的均溫校驗點比較 得到合格的真空爐
for i in range(len(step6_input_excel_df)):
    # 6.5 開始, Excel取出零件的製程溫度需求與真空爐爐號
    #取出零件的溫度需求 str資料
    step6_input_part_temp = step6_input_excel_df.at[i, "製程溫度需求"]
    #取出零件的真空爐爐號 str資料 並用逗號 分開得到列表
    step6_input_part_vaccums = step6_input_excel_df.at[i, "真空爐爐號"].split(",")
    # 6.5 結束
    # 6.6 開始 從零件的真空爐爐號去除故障真空爐 得到零件的製程溫度,無故障真空爐爐號
    # 使用列表推導式過濾出不在 bad_stove 中的元素
    step6_input_part_vaccums = [x for x in step6_input_part_vaccums if x not in bad_stove]
    # 6.6 結束
    # 6.7 開始 對零件的製程溫度進行資料處理
    #零件的溫度需求 str資料 並用逗號分開 同時依照±分開 得到二維列表
    step6_input_part_temp_list = part_temp_split(step6_input_part_temp)
    # 6.7 結束
    # 6.8 開始 零件的所有溫度需求與每個真空爐校驗溫度比較，篩選並紀錄符合條件的爐號
    #傳入 零件的製程溫度、零件可以進入的真空爐爐號(無故障)、真空爐均溫校驗資訊(無故障) 得到該零件經過均溫校驗比較後的真空爐盧號 datatype: list
    step6_part_qualified_vac_list = check_temperature_range(step6_input_part_temp_list,step6_input_part_vaccums,vacuum_TUS_dict)
    #合併多個真空爐爐號，以逗號做間隔
    step6_part_qualified_vac_str = ",".join(check_temperature_range(step6_input_part_temp_list,step6_input_part_vaccums ,vacuum_TUS_dict))
    step6_input_excel_df.at[i, "可進入的真空爐爐號"] = step6_part_qualified_vac_str
    # 6.8 結束
# 6.9 開始
#step6結果轉成Excel檔案
step6_input_excel_df.to_excel(os.path.join(script_dir, '程式間接輸出/step6_output_excel.xlsx'), index=False)
# 6.9 結束
#--------------------------------------step6----------------------------------------
#--------------------------------------step 7 ----------------------------------------
def Step7_scheduling():
    # 欲使用的Excel檔。
    # 7.1 開始,輸入Step 6 Excel結果
    buffer = [os.path.join(script_dir, '程式間接輸出/step6_output_excel.xlsx')]
    # 回傳三個變數 -> 儲存Excel的變數 , 儲存Excel所有sheet名稱的變數 , 儲存Excel第一個(最左)sheet的變數
    workbook_buffer, sheet_names_buffer, sheet_buffer = excel_init_setup(buffer)
    # 新增的進入爐子的欄位的位置
    target_col = sheet_buffer[0].max_column + 1
    # 命名欄位名稱
    sheet_buffer[0].cell(1, target_col).value = "最後進入的真空爐"
    # 7.1 結束
    # 7.3 開始,計算未安排真空爐的佔爐率 = 零件數量/真空爐數量限制 並記錄滿爐、80%以上、80%以下真空爐
    # 使用到的欄位以及相應excel位置
    tag_name = ["數量", "A0064232(32001749)數量限制", "A0041090(32001661)數量限制", "32003075-1數量限制", "32013259數量限制", "32018276數量限制", "32022219數量限制", "32025349數量限制", "32025350數量限制", "可進入的真空爐爐號"]
    tag_name_2 =["A0064232(32001749)", "A0041090(32001661)", "32003075-1", "32013259", "32018276", "32022219", "32025349", "32025350"]
    tag_pos = []
    tag_dict = {}
    # print("在sheet的指定工作表中透過tag找到其對應的row,col位置")
    # 創建字典 key是真空爐數量限制以及數量、真空爐爐號  字典的value是 Step6輸出的Excel key欄位名稱的column位置 ，
    # 裡面真空爐數量限制與真空険爐號col對應 例如: 字典中 key = A0064232(32001749)數量限制 value = col 2 而key = A0064232(32001749) value = col 2
    for i in range(len(tag_name)):
        tag_pos.append(get_tag_position(sheet_buffer[0], tag_name[i]))
        # print(f"\n例如要找 {tag_name[i]} , 他會在(row,col) = {tag_pos[i]}")
        tag_dict[f"{tag_name[i]}"] = tag_pos[i][1]
        # print("接著在字典中設定他對應的col位置:")
        # print(f"dict['{tag_name[i]}'] = {tag_pos[i][1]}")
        # 將每個爐子的爐號與每個爐子的數量限制 設置相同的col位置。
        if tag_name[i] != "數量" and tag_name[i] != "可進入的真空爐爐號":
            print(f"\n如果 {tag_name[i]} -> {tag_pos[i][1]} , 那麼 {tag_name_2[i-1]} 同樣會對應到 {tag_pos[i][1]} 這個位置,如: dict['{tag_name_2[i-1]}'] = {tag_pos[i][1]}")
            tag_dict[f"{tag_name_2[i-1]}"] = tag_pos[i][1]
    # 顯示每個欄位對應的位置(row,col)
    for i in range(len(tag_pos)):
        print(f"{tag_pos[i]} - {tag_name[i]}")
    # 7.3 結束
    # 針對每列(各個零件)開始處理
    for i in range(2, sheet_buffer[0].max_row+1):

        # 7.2 開始,取出零件的數量與可進入真空爐爐號同時初始化真空爐狀態
        # 初始化變數----------------------------
        # 用來判定是否脫離迴圈
        break_pause = True
        # 欲填入新欄位的字串(選擇的真空爐)
        # 用來查看該爐是否已排程 0代表還沒排過 1代表已經排過
        enter_stove = [1] * 8
        string_buffer = ""
        # 查看當列是否存在，不存在則代表資料有誤，暫停程式。
        if sheet_buffer[0].cell(i, 1).value == None:
            print(f"this row - ({i} , 1) is empty...")
            exit(1)
        # 數量 , 可進入的真空盧
        # 零件的數量  從Step6輸出Excel的數量欄位擷取       # 可進入的真空爐為一個列表，例如分別儲存A0064232(32001749),A0041090(32001661),32003075-1,32013259,32018276,32022219,32025349,32025350...
        count = sheet_buffer[0].cell(i, tag_dict["數量"]).value
        # 零件的可進入真空爐 從Step6輸出Excel的可進入真空爐 欄位擷取
        # 7/23新增 如果發生 經過step6篩選後零件沒有真空爐可以進入加工，則填入"沒有可以進入的真空爐"，並跳過
        if sheet_buffer[0].cell(i, tag_dict["可進入的真空爐爐號"]).value == None:
            sheet_buffer[0].cell(i, target_col).value = "沒有可以進入的真空爐"
            continue
        ok_stove = sheet_buffer[0].cell(i, tag_dict["可進入的真空爐爐號"]).value.split(',')
        # 將可以進入的真空盧查看表(enter_stove)設置 根據可進入的真空爐 把1依照對應位置設為0 ，若此階段不是都為0 同樣代表零件本身不能進入那個真空爐 為1
        for j in ok_stove:
            enter_stove[tag_dict[j] - 9] = 0
        # 7.2 結束
        # 先查看每個爐子是否已排程，如果沒有才往下做
        # 先判斷該零件數量是否大於數量限制，沒有則先計算滿爐率
        # 在判斷的同時，會以100%為優先，如果未滿100%，則只會抓大於80%的最大滿爐率的位置
        # 對一個爐子排程完後，會再對那個零件的所有爐子在找一輪 看有沒有滿80%獲滿100%的爐子，不用擔心找到重複，因為entry_stove設為1 代表那個爐子已找過
        #全部爐子都變成1 代表爐子都排成或找過 就進入到分支選項並將break_pause = False 同時輸入結果
        # 設定隨機數生成器的種子 為了每次執行能出現固定的結果
        seed_value = 1
        random.seed(seed_value)
        while break_pause == True:
            # 初始化
            max_80_max_buffer = False
            max_100_max_buffer = False
            papa_80_list = []
            papa_100_list = []
            # 開始對零件的每個可行真空爐計算佔爐率，紀錄 一個零件中 哪些是100% 哪些有80%以上
            #其結果會出現 要只有一個或多個100%爐子其餘都是80% 或者都是80%爐子
            # 7.3 開始,計算未安排真空爐的佔爐率 = 零件數量/真空爐數量限制 並記錄滿爐、80%以上、80%以下真空爐
            for j in ok_stove:
                # 如果數量小於0，會直接退出迴圈判斷
                if count <= 0 :
                    break
                # 如果當前真空爐為1，代表此爐已安排，或是無法排程，也就是檢查過了。減9 因為step6輸出的Excel 真空爐1號數量限制在column 9 相減為0 代表enter_stove的第一個爐子，
                if enter_stove[tag_dict[j] - 9] == 0:
                    # 如果零件的數量是大於數量限制，那就代表其為100%滿爐。
                    if count >= sheet_buffer[0].cell(i, tag_dict[j]).value:
                        # 找到了已擁有100%滿爐的爐子
                        max_100_max_buffer = True
                        papa_100_list.append(j)
                    else:
                        # 如果已經有了100%滿爐的爐子，那就不需要去檢查未滿100%的爐子了。
                        if max_100_max_buffer == False:
                            #沒有100%滿爐 則計算爐子的數量占比
                            temp_value = (count / sheet_buffer[0].cell(i, tag_dict[j]).value) * 100
                            # 超過80% 則可將此零件安排到此真空爐
                            if temp_value > 80:
                                max_80_max_buffer = True
                                papa_80_list.append(j)
            # 7.3 結束
            # 比對過一輪之後先判斷是否有100%滿爐的
            # 7.4 開始,是否有滿爐(佔爐率100%),將零件隨機分配到任一符合條件的真空爐分配結果以"真空爐爐號=數量限制"儲存並把安排過的真空爐紀錄為已安排零件數量減去數量限制
            if max_100_max_buffer == True:
                # 隨機抽一個爐子排程
                random_stove = random.choice(papa_100_list)
                string_buffer = string_buffer + random_stove + '=' + str(sheet_buffer[0].cell(i, tag_dict[random_stove]).value) + ','
                # 將抽到的該爐設為排程過的狀態
                enter_stove[tag_dict[random_stove] - 9] = 1
                # 零件數量扣除已排入的數量限制
                count -= sheet_buffer[0].cell(i, tag_dict[random_stove]).value # count = count - sheet_buffer[0].cell(i, tag_dict[j]).value
            # 7.4 結束
            # 7.5 開始,是否有佔爐率80%以上真空爐,將零件隨機分配到任一符合條件的真空爐分配結果以"真空爐爐號=零件數量"儲存,並把安排過的真空爐紀錄為已安排,零件數量設為0判斷是否有滿80%的, 代表都是滿80%的爐子，沒有100%的爐子
            elif max_80_max_buffer == True:
                # 隨機抽一個爐子排程
                random_stove = random.choice(papa_80_list)
                string_buffer = string_buffer + random_stove + '=' + f"{count}" + ','
                # 將抽到的該爐設為排程過的狀態
                enter_stove[tag_dict[random_stove] - 9] = 1
                # 零件數量扣除已排入的數量限制
                count = 0 # count = count - sheet_buffer[0].cell(i, tag_dict[j]).value
                # 7.5 結束
            # 7.6 開始,之前的真空爐是否有分配零件
            # 如果既沒有100%也沒有80%，則退出迴圈並找下一個,同時 如果entry_stove都=1 也會跳過for j in ok_stove迴圈(找真空爐看是否滿爐) 進入到此分支
            else:
                break_pause = False
                # 將成果輸入到新的欄位 如果前面已經有安排零件
                if string_buffer != "":
                    sheet_buffer[0].cell(i, target_col).value = string_buffer.strip(',')
                #如果完全都沒有安排零件 則填入該資訊
                else:
                    sheet_buffer[0].cell(i, target_col).value = "沒有可以進入的真空爐"
            # Step7.6 結束
    # Step7.7 開始,輸出Step 7 Excel結果
    workbook_buffer[0].save(os.path.join(script_dir, '程式間接輸出/step7_output_excel.xlsx'))
    # Step7.7 結束
Step7_scheduling()
#--------------------------------------step 7 ----------------------------------------
#--------------------------------------step 8 ----------------------------------------
#--------------------------------------step 8 函式區 ----------------------------------------
#8.4 開始,創建熱處理排程表
# 建立8個真空爐的dataframe 的字典 key = 爐名 value = dataframe
def Create_stove_df(name_list):
    stove_df_dict = {}
    schedule_columns = ["項次","件號","工作號碼","數量","類別","時間","預計進爐時間","預計出爐時間"]
    for i in name_list:
        stove_df_dict[i] = pd.DataFrame(columns=schedule_columns)
    return stove_df_dict
# 8.4 結束
#8.5 開始,從Step 7 Excel取出"最後進入的真空爐"並對"真空爐爐號=加工數量"進行資料處理
#傳入一個資訊 數個真空爐=數量的字串 輸出二維列表 列表中的每個元素皆為列表[真空爐,數量] 分割 多個爐號和數量
def stove_num_split(schedule_num):
    stove_num_split_list =[]
    #先分割 爐子-數量
    stove_num_list = schedule_num.split(",")
    #在對每個 爐子-數量 依照 - 去分割
    for i in stove_num_list:
        i = i.split("=")
        stove_num_split_list.append(i)
    #回傳二維列表
    return stove_num_split_list
# 8.5 結束
# 8.7 8.10 8.12都會用到
# 輸入當前日期，可輸出當前日期的下周一早上8點
def get_next_monday_8am30min(current_date):
    # 計算下一周一的日期
    next_monday = current_date + pd.DateOffset(days=(7 - current_date.weekday()))
    # 設置為早上8點30分
    next_monday = next_monday.replace(hour=8, minute=30, second=0, microsecond=0)
    return next_monday
def round_up_minute(dt):
    """將分鐘進位到下一個整點或者30分鐘"""
    if dt.minute == 0:
        return dt.replace(second=0, microsecond=0)
    elif dt.minute <= 30:
        minute = 30
    else:
        dt += datetime.timedelta(hours=1)
        minute = 0
    return dt.replace(minute=minute, second=0, microsecond=0)

# 輸入 真空爐排程表(字典 key= 真空爐號碼 value=dataframe零件資料) 真空爐爐號 和 第一個零件的進爐時間 可以將傳進去的真空爐dataframe 去填入每一個零件的預計進爐時間和出爐時間
def update_schedule(stove_df_dict,name,Start_time):
    # Start_time為每一個爐子的第一個零件的預計進爐時間
    # 將真空爐的所有零件 依照加工時間去計算出爐時間 並讓這個出爐時間作為下一個零件的進爐時間 以此類推
    for index, row in stove_df_dict[name].iterrows():
        if index == 0:
            stove_df_dict[name].at[index, "預計進爐時間"] = Start_time
        # 進爐時間 + 工時
        working_hr = row["時間"].replace("hr", "")
        working_hr_int = int(working_hr)
        projected_completion_time = Start_time + pd.Timedelta(hours=working_hr_int)
        # 如果加上工時的出爐時間超過當週六下午五點 則讓零件的預計進爐時間修改至下周一早上8點
        if projected_completion_time > projected_completion_time.replace(hour=17, minute=0, second=0, microsecond=0, day=projected_completion_time.day) - pd.DateOffset(days=(projected_completion_time.weekday() - 5)):
            #pd.DateOffset(days=(projected_completion_time.weekday() - 5)):這個部分是用來計算從當前預計出爐時間到當週六的日期差。
            # weekday() 函数返回的是星期幾的索引（星期一為0，星期日為6），所以 projected_completion_time.weekday() - 5 計算的是從當天到星期六還差幾天。
            # 如果 projected_completion_time 已經是星期六或之後，這個值會是負數或零。
            # 計算出日期差 再將預計出爐時間去扣除這個日期差(日期差都是負數) 自動以那個預計出爐日期去計算當周六下午5點
            # 如果超過，設定進爐時間為下周一早上8點30分
            Start_time = get_next_monday_8am30min(projected_completion_time)
            stove_df_dict[name].at[index, "預計進爐時間"] = Start_time
        #沒超過 正常地將上一個零件的出爐時間作為這一個零件的進爐時間
        else:
            stove_df_dict[name].at[index, "預計進爐時間"] = Start_time
        stove_df_dict[name].at[index, "預計出爐時間"] = Start_time + pd.Timedelta(hours=working_hr_int)
        Start_time = stove_df_dict[name].at[index, "預計出爐時間"]
        #如果這個零件的index是在倒數第二個零件index
        if index < len(stove_df_dict[name]) - 1:
            stove_df_dict[name].at[index + 1, "預計進爐時間"] = Start_time
# 8.7 8.10 8.12都會用到
#--------------------------------------step 8 主程式----------------------------------------
current_timestamp = pd.Timestamp.now().replace(second=0, microsecond=0)
month_day = current_timestamp.strftime('%m-%d')
file_path = f"{schedule_result_filepath}/{month_day}output.xlsx"
# 輸入step7的結果
# 8.1 開始, 輸入Step 7 Excel結果
step7_input_excel = pd.read_excel(os.path.join(script_dir, '程式間接輸出/step7_output_excel.xlsx'), converters={'工作號碼': str})
#8.1結束
#8.2,讀取起始加工時間表
# 輸入起始加工時間設定表
start_timedf = pd.read_excel(start_time_excel_filepath)
# 將起始加工時間設定表 dataframe 每個欄位的資料型態設定成str
start_timedf.columns = start_timedf.columns.map(str)
#8.2 結束
#8.3 開始,Step 7 Excel"最後進入的真空爐"欄位去除"沒有可以進入的真空爐"
# step7 excel結果 只留下排程好的零件資料
Part_processable_filter = (step7_input_excel["最後進入的真空爐"] != "沒有可以進入的真空爐")
Part_processable_df = step7_input_excel[Part_processable_filter].reset_index(drop=True)
# 8.3 結束
stove_name = ["A0064232(32001749)","A0041090(32001661)","32003075-1","32013259","32018276","32022219","32025349","32025350"]
#8.4 開始,創建熱處理排程表
# 創造字典 key=真空爐爐號 value目前是空的
stove_df_dict = Create_stove_df(stove_name)
#8.4 結束
#8.5 開始,從Step 7 Excel取出"最後進入的真空爐"並對"真空爐爐號=加工數量"進行資料處理
#把零件資料根據排程結果寫在Excel中的各個真空爐區塊
for i in range(len(Part_processable_df)):
    schedule_num = Part_processable_df.at[i,"最後進入的真空爐"]
    #得到分割後的二維列表 每個元素都是[爐子,數量]
    schedule_num_split = stove_num_split(schedule_num)
    # 8.5 結束
    # 8.6 開始,依照"真空爐爐號=加工數量"的爐號，找到對應的排程表區域
    # 對每個零件的真空爐排程 填入排程表
    for j in schedule_num_split:
        for name in stove_name:
            if j[0] == name:
                stove_df_dict[name].at[i,"件號"] = Part_processable_df.at[i,"件號"]
                stove_df_dict[name].at[i,"工作號碼"] = Part_processable_df.at[i,"工作號碼"]
                stove_df_dict[name].at[i,"數量"] = j[1]
                stove_df_dict[name].at[i,"類別"] = Part_processable_df.at[i,"類別"]
                stove_df_dict[name].at[i,"時間"] = Part_processable_df.at[i,"時間"]
    # 8.6 結束
    # 8.7 開始,依照起始加工時間表各爐的時間，作為將每個真空爐起始零件進爐時間，接著計算各爐所有零件的進出爐時間
# 將真空爐區塊的零件填寫項次。並根據Start_time填寫第一個零件的預計進爐時間，並依序計算每個零件的進爐和出爐時間
for name in stove_name:
    if stove_df_dict[name].empty:
        continue  # 跳過後續的處理，繼續下一次迴圈
    # 程式將index重製 並且輸入到項次欄位
    stove_df_dict[name].reset_index(drop=True, inplace=True)
    stove_df_dict[name]["項次"] = stove_df_dict[name].index +1
    #先設定每一個爐子的第一個零件的預計進爐時間
    initial_start_time = start_timedf.at[0,name]
    update_schedule(stove_df_dict, name, initial_start_time)
    # 8.7 結束
    # 8.8 開始,是否有輸入前一個排程數據檔
# 如果是第二次以後執行自動排程系統 輸入第一次排程的排程表 6/17新增
if last_schedule_pkl_filepath != "":
    # 8.8 結束
    # 8.9 開始,讀取前一次排程數據與當前執行系統的時間
    # 輸入上一次執行排程系統得到的排程表 pkl檔案 6/17新增
    with open(last_schedule_pkl_filepath, 'rb') as file:
        last_schedule_dict = pickle.load(file)
    # 讀取當前執行系統時間
    current_time = pd.Timestamp.now().replace(second=0, microsecond=0)
    # current_time = pd.Timestamp('2024/8/2 8:30')
    #8.9 結束
    for name in stove_name:
        # 8.10 開始,前一次排程的真空爐是否安排零件
        #如果上一個排程表真空爐沒有加工零件 則去更新這一次排程表真空爐的預計進出爐時間 第一個零件的進爐時間使用當下執行系統的時間
        if last_schedule_dict[name].empty:
            output_time = round_up_minute(current_time)
            update_schedule(stove_df_dict,name,output_time)
            continue  # 跳過後續的處理，繼續下一次迴圈
        # 8.10結束
        # 如果上一個排程表真空爐有零件 第二次執行系統當下真空爐一定有零件正在加工
        processing_no = None
        processing_opno = None
        output_time = None
        drop_indices = []
        # 8.11 開始,前一次排程的真空爐是否全部加工完成
        for index, row in last_schedule_dict[name].iterrows():
            #  Step 8.11開始
            input_stove_time = row["預計進爐時間"]
            output_stove_time = row["預計出爐時間"]
            # 如果上一個排程表 根據第二次執行系統的時間 如果只有已經完成加工的零件 代表那個真空爐完成加工後已經有一段時間沒加工 則把當下執行系統的時間的整點或30分鐘進位作為下一批第一個零件的進爐時間
            # 如果只有正在加工或者同時具有已完成和正在加工的零件 則以正在加工的零件出爐時間作為下一批第一個零件的進爐時間
            #去除已經加工完的零件 第一次排程的排程表預計出爐時間如果小於第二次執行系統的時間 代表當下零件已經加工完成
            if output_stove_time <= current_time:
                drop_indices.append(index)
                # 更新已經完成加工的零件出爐時間 最後會得到當下執行系統的時間 的進位
                output_time = round_up_minute(current_time)
                # 8.11結束
            # 8.12 開始,依照當前時間找到正在加工的零件件號、工作號碼與出爐時間
            #從多個零件中找出一個(只會有一個)正在加工零件的件號、工作號碼、預計出爐時間
            elif input_stove_time <= current_time < output_stove_time:
                processing_no = row["件號"]
                processing_opno = row["工作號碼"]
                output_time = row["預計出爐時間"]
        # last_schedule_dict[name].drop(drop_indices, inplace=True)
        # last_schedule_dict[name].reset_index(drop=True, inplace=True)
        #對真空爐進行更新
        drop_indices = []
        for index, row in stove_df_dict[name].iterrows():
            # 按照件號、工作號碼，找到第二次排程表的加工中零件
            if row["件號"] == processing_no and row["工作號碼"] == processing_opno:
                drop_indices.append(index)
        # 從第二次排程 刪除該加工中零件
        stove_df_dict[name].drop(drop_indices, axis=0, inplace=True)
        stove_df_dict[name].reset_index(drop=True, inplace=True)
        # 重頭更新第二次排程的進出爐時間 第一個零件的進爐時間為加工中零件的預計出爐時間
        update_schedule(stove_df_dict,name,output_time)
        stove_df_dict[name]["項次"] = stove_df_dict[name].index +1
        #  8.12結束
#  8.13 開始 將排程結果存在PKL檔案,並輸出熱處理排程表
# 提取當前的月和日
month_day = current_timestamp.strftime('%m-%d')
with open(f"4-前一天排程數據Pickle/{month_day}stove_df_dict.pkl", 'wb') as file:
    pickle.dump(stove_df_dict, file)

# 找到真空爐中最多的row數量
max_row = max([df.shape[0] for df in stove_df_dict.values()])
# 根據最大列數擴展真空爐的dataframe 為了美觀
for name in stove_name:
    stove_df_dict[name] = stove_df_dict[name].reindex(index=range(max_row), fill_value=pd.NA)

#將8個真空爐的dataframe寫入excel之中
writer = pd.ExcelWriter(file_path ,engine = 'xlsxwriter')
title_row = 0
for name in stove_name:
    stove_df_dict[name].to_excel(writer, index=None,header=True, startrow=title_row + 1)
    #header=True 代表dataframe的欄位名稱也填入
    ws = writer.sheets['Sheet1']
    ws.write_string(title_row, 0, name)
    title_row += len(stove_df_dict[name]) +2
writer.close()
# 8.13 結束
#--------------------------------------step 8 ----------------------------------------

#--------------------------------------使用者輸出介面----------------------------------------
#--------------------------------------使用者輸出介面函式區----------------------------------------
def open_excel():
    global schedule_result_filepath
    #得到當前執行系統時間的月和日
    current_timestamp = pd.Timestamp.now().replace(second=0, microsecond=0)
    month_day = current_timestamp.strftime('%m-%d')
    # 這裡填入你的 幾月幾號熱處理排程表Excel 文件路徑
    excel_path = f"{schedule_result_filepath}/{month_day}output.xlsx"
    print("pos... = " + excel_path)
    # 使用系統默認的應用程式打開 Excel 文件
    run(['start', excel_path], shell=True)

def adjust_label_width(event):
    # 獲取絕對位置的文字
    text = text_label_output_part.cget("text")
    # 獲取絕對位置的長度
    text_length = len(text)
    # 設置 Label 元件的寬度為絕對位置的長度加上一些額外空間
    text_label_output_part.config(width=text_length + 10)
#--------------------------------------使用者輸出介面主程式----------------------------------------
file_path = f"{schedule_result_filepath}/{month_day}output.xlsx"
# 取得當前工作目錄
current_directory = os.getcwd()
# 搜索當前目錄下的output.xlsx
file_pattern = os.path.join(schedule_result_filepath, f'{month_day}output.xlsx')
matching_files = glob.glob(file_pattern)
# 創建主視窗
root = tk.Tk()
root.title("執行成功")
root.minsize(width=500, height=500)
large_font = font.Font(size=15)
# 0731修改 更改了開啟檔案按鈕(改成上面)和文字框(改成下面)的位置
# 創建一個按鈕，點擊時調用 open_excel 函數
button = tk.Button(root, text="打開輸出: 「熱處理排程表單Excel」",font=("標楷體", 15, "bold"), command=open_excel)
button.pack(pady=20)
text_label_start_part = tk.Label(root, text="輸出位置", width=10, font = large_font)
text_label_start_part.pack(pady=20)
if matching_files:
    # 取得第一個匹配文件的絕對路徑
    absolute_path = os.path.abspath(matching_files[0])
    # 顯示輸出的絕對位置
    text_label_output_part = tk.Label(root, text=schedule_result_filepath, width=50, bg='white')
    text_label_output_part.pack(pady=0)
else:
    print("File 'output.xlsx' not found in the current directory.")



# 綁定調整 Label 寬度的函式到視窗大小變化事件
root.bind("<Configure>", adjust_label_width)

# 運行 tkinter 事件循環
root.mainloop()

#-------------------------------------- 使用者輸出介面 ----------------------------------------





